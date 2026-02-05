"""
Excel Submission Evaluator for O-Level Computing (school-telegram-portal).
Evaluates student spreadsheet submissions against an answer key and mark scheme.
Single source of truth for SALES_ANALYSIS marking; used by utils.spreadsheet_evaluator and scripts/evaluate_submissions.py.
"""

import re
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class CellResult:
    """Result for a single cell evaluation"""
    cell_ref: str
    question_num: int
    marks_possible: int
    marks_awarded: float
    student_formula: Optional[str]
    expected_formula: Optional[str]
    student_value: Any
    expected_value: Any
    feedback: str
    formula_correct: bool
    value_correct: bool


@dataclass
class QuestionResult:
    """Result for a question (may span multiple cells)"""
    question_num: int
    description: str
    total_marks: int
    marks_awarded: float
    cells: List[CellResult] = field(default_factory=list)
    feedback: str = ""


@dataclass
class EvaluationResult:
    """Complete evaluation result"""
    student_file: str
    student_name: str
    total_marks: int
    marks_awarded: float
    percentage: float
    questions: List[QuestionResult] = field(default_factory=list)
    summary: str = ""


# ============================================================
# MARK SCHEME
# ============================================================

class MarkScheme:
    """Define the mark scheme for the SALES_ANALYSIS task"""

    def __init__(self):
        self.questions = [
            {
                "num": 1,
                "description": "G4:G15 - Calculate 2025 Total Sales using SUM formula",
                "marks": 1,
                "cells": [f"G{i}" for i in range(4, 16)],
                "expected_formula_pattern": r"SUM\([C-F]\d+:[C-F]\d+\)",
                "formula_type": "SUM",
                "marking_notes": "1 mark for correct SUM formula summing quarterly sales"
            },
            {
                "num": 2,
                "description": "H4:H15 - Calculate Commission using VLOOKUP from Commission Table",
                "marks": 2,
                "cells": [f"H{i}" for i in range(4, 16)],
                "expected_formula_pattern": r"VLOOKUP\(.*\$?A\$?20.*\$?C\$?24.*3.*TRUE\).*\*",
                "formula_type": "VLOOKUP",
                "marking_notes": "1 mark for VLOOKUP function with correct range, 1 mark for multiplying by sales"
            },
            {
                "num": 3,
                "description": "J4:J15 - IF function to show Exceed/Miss based on target comparison",
                "marks": 2,
                "cells": [f"J{i}" for i in range(4, 16)],
                "expected_formula_pattern": r'IF\(G\d+>I\d+,.*EXCEED.*,.*MISS',
                "formula_type": "IF",
                "marking_notes": "1 mark for IF function, 1 mark for correct comparison logic"
            },
            {
                "num": 4,
                "description": "Conditional formatting - Red background for rows where target is missed",
                "marks": 3,
                "cells": ["A4:J15"],
                "expected_formula_pattern": None,
                "formula_type": "CONDITIONAL_FORMATTING",
                "marking_notes": "1 mark for applying to correct range, 1 mark for correct condition, 1 mark for red fill"
            },
            {
                "num": 5,
                "description": "I19:I22 - SUMIF for departmental sales totals",
                "marks": 4,
                "cells": ["I19", "I20", "I21", "I22"],
                "expected_formula_pattern": r"SUMIF\(.*B.*:.*B.*,.*,.*G.*:.*G.*\)",
                "formula_type": "SUMIF",
                "marking_notes": "1 mark per department (Home Appliances, IT, Electrical, Beddings)"
            },
            {
                "num": 6,
                "description": "I23 - SUM function for Company Total Sales",
                "marks": 1,
                "cells": ["I23"],
                "expected_formula_pattern": r"SUM\(I19:I22\)",
                "formula_type": "SUM",
                "marking_notes": "1 mark for SUM of department totals"
            },
            {
                "num": 7,
                "description": "I25 - Formula for Sales yet to be achieved",
                "marks": 1,
                "cells": ["I25"],
                "expected_formula_pattern": r"I24-I23",
                "formula_type": "SUBTRACTION",
                "marking_notes": "1 mark for correct subtraction (Target - Current Sales)"
            },
            {
                "num": 8,
                "description": "I26 - Calculate days remaining until end of year",
                "marks": 1,
                "cells": ["I26"],
                "expected_formula_pattern": r"(DATEDIF|DAYS|H25-H23|DATE|82)",
                "formula_type": "DATE_CALCULATION",
                "marking_notes": "1 mark for calculating days between 10/10/2025 and 31/12/2025 (should be 82)"
            }
        ]

    def get_question(self, num: int) -> Optional[Dict]:
        for q in self.questions:
            if q["num"] == num:
                return q
        return None

    def get_total_marks(self) -> int:
        return sum(q["marks"] for q in self.questions)


# ============================================================
# EVALUATOR
# ============================================================

class ExcelEvaluator:
    """Main evaluation class"""

    def __init__(self, answer_key_path: str, mark_scheme: MarkScheme = None):
        self.answer_key_path = answer_key_path
        self.mark_scheme = mark_scheme or MarkScheme()

        self.wb_ans = openpyxl.load_workbook(answer_key_path, data_only=False)
        self.ws_ans = self.wb_ans.active

        self.wb_ans_values = openpyxl.load_workbook(answer_key_path, data_only=True)
        self.ws_ans_values = self.wb_ans_values.active

    def normalize_formula(self, formula: str) -> str:
        if not formula:
            return ""
        f = formula.strip()
        if f.startswith("="):
            f = f[1:]
        f = f.upper()
        f = re.sub(r'\s+', '', f)
        return f

    def check_formula_pattern(self, formula: str, pattern: str) -> bool:
        if not formula or not pattern:
            return False
        normalized = self.normalize_formula(formula)
        try:
            return bool(re.search(pattern, normalized, re.IGNORECASE))
        except Exception:
            return False

    def compare_values(self, student_val: Any, expected_val: Any, tolerance: float = 0.01) -> bool:
        if student_val is None and expected_val is None:
            return True
        if student_val is None or expected_val is None:
            return False

        if isinstance(student_val, str) and isinstance(expected_val, str):
            return student_val.strip().lower() == expected_val.strip().lower()

        try:
            s_num = float(student_val)
            e_num = float(expected_val)
            if e_num == 0:
                return abs(s_num) < tolerance
            return abs(s_num - e_num) / abs(e_num) < tolerance
        except (ValueError, TypeError):
            pass

        return str(student_val).strip().lower() == str(expected_val).strip().lower()

    def evaluate_cell(self, ws_student, ws_student_values, cell_ref: str,
                      question: Dict) -> CellResult:
        student_cell = ws_student[cell_ref]
        student_value_cell = ws_student_values[cell_ref]
        ans_cell = self.ws_ans[cell_ref]
        ans_value_cell = self.ws_ans_values[cell_ref]

        student_formula = student_cell.value if isinstance(student_cell.value, str) and student_cell.value.startswith('=') else None
        expected_formula = ans_cell.value if isinstance(ans_cell.value, str) and ans_cell.value.startswith('=') else None

        student_value = student_value_cell.value
        expected_value = ans_value_cell.value

        formula_correct = False
        if question["expected_formula_pattern"]:
            if student_formula:
                formula_correct = self.check_formula_pattern(
                    student_formula,
                    question["expected_formula_pattern"]
                )

        value_correct = self.compare_values(student_value, expected_value)

        feedback_parts = []
        if not student_formula:
            feedback_parts.append("No formula entered")
        elif not formula_correct:
            feedback_parts.append(f"Formula structure incorrect. Expected pattern using {question['formula_type']}")

        if not value_correct:
            feedback_parts.append(f"Value incorrect. Expected: {expected_value}, Got: {student_value}")

        if formula_correct and value_correct:
            feedback_parts.append("Correct!")

        return CellResult(
            cell_ref=cell_ref,
            question_num=question["num"],
            marks_possible=0,
            marks_awarded=0,
            student_formula=student_formula,
            expected_formula=expected_formula,
            student_value=student_value,
            expected_value=expected_value,
            feedback=" ".join(feedback_parts),
            formula_correct=formula_correct,
            value_correct=value_correct
        )

    def evaluate_conditional_formatting(self, ws_student, question: Dict) -> QuestionResult:
        result = QuestionResult(
            question_num=question["num"],
            description=question["description"],
            total_marks=question["marks"],
            marks_awarded=0
        )

        feedback_parts = []
        marks = 0

        cf_rules = getattr(ws_student.conditional_formatting, '_cf_rules', {}) or {}

        if not cf_rules:
            feedback_parts.append("No conditional formatting found.")
            result.feedback = " ".join(feedback_parts)
            return result

        found_correct_range = False
        found_correct_formula = False
        found_red_fill = False

        for range_string, rules in cf_rules.items():
            range_str = str(range_string)

            if any(x in range_str.upper() for x in ['A4', 'J15', 'A4:J15']):
                found_correct_range = True

            for rule in rules:
                if hasattr(rule, 'formula') and rule.formula:
                    formula_str = str(rule.formula).upper()
                    if 'MISS' in formula_str or ('J' in formula_str and ('=' in formula_str or 'IF' in formula_str)):
                        found_correct_formula = True

                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fill = rule.dxf.fill
                    if hasattr(fill, 'bgColor') and fill.bgColor:
                        color = str(fill.bgColor.rgb).upper() if fill.bgColor.rgb else ""
                        if color and len(color) >= 6:
                            if len(color) == 8:
                                r, g, b = int(color[2:4], 16), int(color[4:6], 16), int(color[6:8], 16)
                            else:
                                r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                            if r > 200 and g < 150 and b < 150:
                                found_red_fill = True

        if found_correct_range:
            marks += 1
            feedback_parts.append("✓ Correct range applied.")
        else:
            feedback_parts.append("✗ Range should cover rows 4-15.")

        if found_correct_formula:
            marks += 1
            feedback_parts.append("✓ Condition formula correctly checks for 'Miss'.")
        else:
            feedback_parts.append("✗ Formula should check if column J = 'Miss'.")

        if found_red_fill:
            marks += 1
            feedback_parts.append("✓ Red background fill applied.")
        else:
            feedback_parts.append("✗ Fill color should be red.")

        result.marks_awarded = marks
        result.feedback = " ".join(feedback_parts)
        return result

    def evaluate_question(self, ws_student, ws_student_values, question: Dict) -> QuestionResult:
        if question["formula_type"] == "CONDITIONAL_FORMATTING":
            return self.evaluate_conditional_formatting(ws_student, question)

        result = QuestionResult(
            question_num=question["num"],
            description=question["description"],
            total_marks=question["marks"],
            marks_awarded=0
        )

        for cell_ref in question["cells"]:
            cell_result = self.evaluate_cell(ws_student, ws_student_values, cell_ref, question)
            result.cells.append(cell_result)

        if question["num"] in [1, 6, 7, 8]:
            correct_cells = sum(1 for c in result.cells if c.formula_correct or c.value_correct)
            if correct_cells == len(result.cells):
                result.marks_awarded = question["marks"]
            elif correct_cells > 0:
                result.marks_awarded = 0.5

        elif question["num"] == 2:
            has_vlookup = any('VLOOKUP' in (c.student_formula or '').upper() for c in result.cells)
            has_multiply = any('*' in (c.student_formula or '') for c in result.cells)
            correct_values = sum(1 for c in result.cells if c.value_correct)

            if has_vlookup:
                result.marks_awarded += 1
            if has_multiply and correct_values > len(result.cells) // 2:
                result.marks_awarded += 1

        elif question["num"] == 3:
            has_if = any('IF(' in (c.student_formula or '').upper() for c in result.cells)
            correct_values = sum(1 for c in result.cells if c.value_correct)

            if has_if:
                result.marks_awarded += 1
            if correct_values == len(result.cells):
                result.marks_awarded += 1
            elif correct_values > len(result.cells) // 2:
                result.marks_awarded += 0.5

        elif question["num"] == 5:
            correct_cells = sum(1 for c in result.cells if c.formula_correct or c.value_correct)
            result.marks_awarded = correct_cells

        correct_count = sum(1 for c in result.cells if c.formula_correct and c.value_correct)
        total_count = len(result.cells)

        feedback_parts = [f"{correct_count}/{total_count} cells correct."]

        incorrect_cells = [c for c in result.cells if not (c.formula_correct and c.value_correct)]
        if incorrect_cells and len(incorrect_cells) <= 3:
            for c in incorrect_cells:
                feedback_parts.append(f"{c.cell_ref}: {c.feedback}")
        elif incorrect_cells:
            feedback_parts.append(f"Check cells: {', '.join(c.cell_ref for c in incorrect_cells[:5])}")

        result.feedback = " ".join(feedback_parts)
        return result

    def extract_student_name(self, filepath: str) -> str:
        filename = Path(filepath).stem
        match = re.search(r'SALES_ANALYSIS_(.+?)_\d+', filename, re.IGNORECASE)
        if match:
            return match.group(1).replace('_', ' ').title()

        clean_name = filename.replace('SALES_ANALYSIS_', '').replace('_', ' ')
        return clean_name if clean_name else "Unknown Student"

    def evaluate(self, student_file_path: str) -> EvaluationResult:
        wb_student = openpyxl.load_workbook(student_file_path, data_only=False)
        ws_student = wb_student.active

        wb_student_values = openpyxl.load_workbook(student_file_path, data_only=True)
        ws_student_values = wb_student_values.active

        student_name = self.extract_student_name(student_file_path)

        result = EvaluationResult(
            student_file=student_file_path,
            student_name=student_name,
            total_marks=self.mark_scheme.get_total_marks(),
            marks_awarded=0,
            percentage=0
        )

        for question in self.mark_scheme.questions:
            q_result = self.evaluate_question(ws_student, ws_student_values, question)
            result.questions.append(q_result)
            result.marks_awarded += q_result.marks_awarded

        result.percentage = (result.marks_awarded / result.total_marks) * 100 if result.total_marks > 0 else 0

        summary_parts = [
            f"Student: {student_name}",
            f"Total Score: {result.marks_awarded}/{result.total_marks} ({result.percentage:.1f}%)",
            "",
            "Question Breakdown:"
        ]

        for q in result.questions:
            status = "✓" if q.marks_awarded == q.total_marks else "△" if q.marks_awarded > 0 else "✗"
            summary_parts.append(f"  Q{q.question_num}: {q.marks_awarded}/{q.total_marks} {status}")

        result.summary = "\n".join(summary_parts)

        return result


# ============================================================
# REPORT GENERATORS (EvaluationResult-based)
# ============================================================

def generate_text_report(result: EvaluationResult) -> str:
    """Generate a detailed text feedback report from an EvaluationResult."""
    lines = [
        "=" * 60,
        "EXCEL EVALUATION REPORT",
        "=" * 60,
        f"Student: {result.student_name}",
        f"File: {Path(result.student_file).name}",
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"TOTAL SCORE: {result.marks_awarded}/{result.total_marks} ({result.percentage:.1f}%)",
        "",
        "-" * 60,
        "QUESTION BREAKDOWN",
        "-" * 60,
    ]

    for q in result.questions:
        status = "✓" if q.marks_awarded == q.total_marks else "△" if q.marks_awarded > 0 else "✗"
        lines.append(f"Q{q.question_num}: {q.marks_awarded}/{q.total_marks} {status} - {q.description}")

    lines.extend(["", "-" * 60, "DETAILED FEEDBACK", "-" * 60])

    for q in result.questions:
        lines.append("")
        lines.append(f"Question {q.question_num}: {q.description}")
        lines.append(f"Marks: {q.marks_awarded}/{q.total_marks}")
        lines.append(f"Feedback: {q.feedback}")

        if q.cells:
            incorrect_cells = [c for c in q.cells if not (c.formula_correct and c.value_correct)]
            if incorrect_cells:
                lines.append("Issues found:")
                for cell in incorrect_cells[:5]:
                    lines.append(f"  • {cell.cell_ref}: {cell.feedback}")
                    if cell.student_formula:
                        lines.append(f"      Your formula: {cell.student_formula}")

    lines.extend(["", "=" * 60, "END OF REPORT", "=" * 60])

    return "\n".join(lines)


def generate_excel_report(results: List[EvaluationResult], output_path: str) -> None:
    """Generate an Excel summary report for multiple students."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marking Summary"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["Student Name", "File", "Total Marks", "Percentage"]
    if results:
        for q in results[0].questions:
            headers.append(f"Q{q.question_num}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_num, result in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=result.student_name).border = border
        ws.cell(row=row_num, column=2, value=Path(result.student_file).name).border = border

        marks_cell = ws.cell(row=row_num, column=3, value=f"{result.marks_awarded}/{result.total_marks}")
        marks_cell.border = border
        marks_cell.alignment = Alignment(horizontal='center')

        pct_cell = ws.cell(row=row_num, column=4, value=f"{result.percentage:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = Alignment(horizontal='center')
        pct_cell.fill = pass_fill if result.percentage >= 50 else fail_fill

        for col, q in enumerate(result.questions, 5):
            q_cell = ws.cell(row=row_num, column=col, value=f"{q.marks_awarded}/{q.total_marks}")
            q_cell.border = border
            q_cell.alignment = Alignment(horizontal='center')
            if q.marks_awarded == q.total_marks:
                q_cell.fill = pass_fill
            elif q.marks_awarded == 0:
                q_cell.fill = fail_fill

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    wb.save(output_path)


def generate_json_report(results: List[EvaluationResult], output_path: str) -> None:
    """Generate JSON report for integration with other systems."""
    import json
    data = {
        "generated_at": datetime.now().isoformat(),
        "total_students": len(results),
        "results": []
    }

    for result in results:
        student_data = {
            "student_name": result.student_name,
            "file": result.student_file,
            "total_marks": result.total_marks,
            "marks_awarded": result.marks_awarded,
            "percentage": result.percentage,
            "questions": []
        }

        for q in result.questions:
            q_data = {
                "question_num": q.question_num,
                "description": q.description,
                "total_marks": q.total_marks,
                "marks_awarded": q.marks_awarded,
                "feedback": q.feedback
            }
            student_data["questions"].append(q_data)

        data["results"].append(student_data)

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)
